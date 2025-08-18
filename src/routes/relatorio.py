from flask import Blueprint, Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime
from src.models.produto import Produto
from src.models.contagem import Contagem
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

relatorio_bp = Blueprint('relatorio', __name__)

@relatorio_bp.route('/api/relatorio/pdf', methods=['GET'])
def gerar_relatorio_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    titulo = Paragraph("<b>Relatório de Estoque</b>", styles['Title'])
    subtitulo = Paragraph("Relatório de Estoque - Detalhado por Lote", styles['Heading2'])
    data_emissao = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])

    elements.append(titulo)
    elements.append(data_emissao)
    elements.append(Spacer(1, 12))
    elements.append(subtitulo)
    elements.append(Spacer(1, 12))

    # Cabeçalho da tabela (sem coluna Cadastro)
    table_data = [["Código", "Nome do Produto", "Lote", "Validade", "Qtd"]]

    produtos = Produto.query.all()

    for produto in produtos:
        contagens = Contagem.query.filter_by(produto_codigo=produto.codigo).all()
        subtotal = sum(c.quantidade for c in contagens)

        for c in contagens:
            validade = f"{c.validade_mes:02d}/{c.validade_ano}" if c.validade_mes and c.validade_ano else "-"
            table_data.append([
                str(produto.codigo),
                produto.nome,
                c.lote,
                validade,
                str(c.quantidade)
            ])

        # Linha subtotal destacada (somente o código)
        table_data.append([
            str(produto.codigo),
            "",
            "",
            "Subtotal",
            str(subtotal)
        ])

    # Monta tabela
    table = Table(table_data, colWidths=[70, 200, 80, 100, 60])
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("ALIGN", (0,1), (0,-1), "LEFT"),
        ("ALIGN", (-1,1), (-1,-1), "RIGHT"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
    ]))

    # Destacar todas as linhas de Subtotal
    for i, row in enumerate(table_data):
        if row[3] == "Subtotal":
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,i), (-1,i), colors.lightyellow),
                ("FONTNAME", (0,i), (-1,i), "Helvetica-Bold"),
                ("ALIGN", (0,i), (0,i), "LEFT"),
            ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = Response(buffer.read(), mimetype='application/pdf')
    response.headers['Content-Disposition'] = f'inline; filename=relatorio_estoque_{datetime.now().date()}.pdf'
    return response


@relatorio_bp.route('/api/relatorio/excel', methods=['GET'])
def gerar_relatorio_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Estoque"

    # Cabeçalho sem a coluna Cadastro
    headers = ["Código", "Nome do Produto", "Lote", "Validade", "Qtd"]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"].font = Font(bold=True)
        ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")

    produtos = Produto.query.all()

    for produto in produtos:
        contagens = Contagem.query.filter_by(produto_codigo=produto.codigo).all()
        subtotal = sum(c.quantidade for c in contagens)

        for c in contagens:
            validade = f"{c.validade_mes:02d}/{c.validade_ano}" if c.validade_mes and c.validade_ano else "-"
            ws.append([
                str(produto.codigo),
                produto.nome,
                c.lote,
                validade,
                c.quantidade
            ])

        # Linha subtotal destacada (somente o código)
        row_idx = ws.max_row + 1
        ws.append([
            str(produto.codigo),
            "",
            "",
            "Subtotal",
            subtotal
        ])
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = Response(buffer.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_estoque_{datetime.now().date()}.xlsx'
    return response
