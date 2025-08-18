from flask import Blueprint, jsonify, send_file
from src.models.user import db
from src.models.produto import Produto
from src.models.contagem import Contagem
from sqlalchemy import func
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import os
from datetime import datetime

relatorio_novo_bp = Blueprint('relatorio_novo', __name__)

@relatorio_novo_bp.route('/relatorio/pdf_novo', methods=['GET'])
def gerar_relatorio_pdf_novo():
    """Gera relatório em PDF no novo formato"""
    try:
        # Buscar todos os produtos com suas contagens (incluindo produtos sem contagem)
        produtos_data = db.session.query(
            Produto.codigo,
            Produto.nome
        ).order_by(Produto.codigo).all()
        
        # Buscar todas as contagens
        contagens_data = db.session.query(
            Produto.codigo,
            Produto.nome,
            Contagem.lote,
            Contagem.validade_mes,
            Contagem.validade_ano,
            Contagem.quantidade
        ).join(Produto).order_by(Produto.codigo, Contagem.lote).all()
        
        # Criar PDF em memória
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,  # Center
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1,  # Center
            fontName='Helvetica'
        )
        
        # Elementos do documento
        elements = []
        
        # Título
        title = Paragraph("Relatório de Estoque", title_style)
        elements.append(title)
        
        # Data de geração
        data_geracao = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style)
        elements.append(data_geracao)
        elements.append(Spacer(1, 20))
        
        # Subtítulo
        subtitle = Paragraph("Relatório de Estoque - Detalhado por Lote", ParagraphStyle(
            'CustomSubtitle2',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            fontName='Helvetica-Bold'
        ))
        elements.append(subtitle)
        
        # Tabela principal
        table_data = [['Código', 'Nome do Produto', 'Lote', 'Validade', 'Qtd', 'Cadastro']]
        
        meses = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        # Organizar dados por produto
        produtos_dict = {}
        for produto in produtos_data:
            produtos_dict[produto.codigo] = {
                'nome': produto.nome,
                'contagens': [],
                'subtotal': 0
            }
        
        # Adicionar contagens aos produtos
        for contagem in contagens_data:
            if contagem.codigo in produtos_dict:
                validade = f"{contagem.validade_mes:02d}/{contagem.validade_ano}"
                cadastro = '-'  # Campo não disponível no modelo atual
                
                produtos_dict[contagem.codigo]['contagens'].append({
                    'lote': contagem.lote,
                    'validade': validade,
                    'quantidade': contagem.quantidade,
                    'cadastro': cadastro
                })
                produtos_dict[contagem.codigo]['subtotal'] += contagem.quantidade
        
        total_geral = 0
        
        # Adicionar dados à tabela
        for codigo in sorted(produtos_dict.keys()):
            produto = produtos_dict[codigo]
            
            if produto['contagens']:
                # Produto com contagens
                for i, contagem in enumerate(produto['contagens']):
                    if i == 0:
                        # Primeira linha do produto
                        table_data.append([
                            codigo,
                            produto['nome'],
                            contagem['lote'],
                            contagem['validade'],
                            str(contagem['quantidade']),
                            contagem['cadastro']
                        ])
                    else:
                        # Linhas subsequentes (sem repetir código e nome)
                        table_data.append([
                            codigo,
                            produto['nome'],
                            contagem['lote'],
                            contagem['validade'],
                            str(contagem['quantidade']),
                            contagem['cadastro']
                        ])
                
                # Linha de subtotal
                table_data.append([
                    f'Subtotal {codigo}:',
                    '',
                    '',
                    '',
                    str(produto['subtotal']),
                    ''
                ])
                total_geral += produto['subtotal']
            else:
                # Produto sem contagens (estoque zerado)
                table_data.append([
                    codigo,
                    produto['nome'],
                    '-',
                    '-',
                    '0',
                    '-'
                ])
                # Linha de subtotal para produto zerado
                table_data.append([
                    f'Subtotal {codigo}:',
                    '',
                    '',
                    '',
                    '0',
                    ''
                ])
        
        # Linha de total geral
        table_data.append([
            'TOTAL GERAL:',
            '',
            '',
            '',
            str(total_geral),
            ''
        ])
        
        # Criar tabela
        table = Table(table_data, colWidths=[0.8*inch, 2.2*inch, 0.8*inch, 1*inch, 0.8*inch, 1*inch])
        
        # Estilo da tabela
        table_style = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo da tabela
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Total geral
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]
        
        # Destacar linhas de subtotal
        row_count = 1  # Começar após o cabeçalho
        for codigo in sorted(produtos_dict.keys()):
            produto = produtos_dict[codigo]
            if produto['contagens']:
                row_count += len(produto['contagens'])
            else:
                row_count += 1
            
            # Linha de subtotal
            table_style.append(('BACKGROUND', (0, row_count), (-1, row_count), colors.lightblue))
            table_style.append(('FONTNAME', (0, row_count), (-1, row_count), 'Helvetica-Bold'))
            row_count += 1
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
        
        # Número da página
        elements.append(Spacer(1, 20))
        page_num = Paragraph("Página 1", ParagraphStyle(
            'PageNum',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1  # Center
        ))
        elements.append(page_num)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y-%m-%d")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar PDF: {str(e)}'}), 500

@relatorio_novo_bp.route('/relatorio/excel_novo', methods=['GET'])
def gerar_relatorio_excel_novo():
    """Gera relatório em Excel no novo formato"""
    try:
        # Buscar todos os produtos com suas contagens (incluindo produtos sem contagem)
        produtos_data = db.session.query(
            Produto.codigo,
            Produto.nome
        ).order_by(Produto.codigo).all()
        
        # Buscar todas as contagens
        contagens_data = db.session.query(
            Produto.codigo,
            Produto.nome,
            Contagem.lote,
            Contagem.validade_mes,
            Contagem.validade_ano,
            Contagem.quantidade
        ).join(Produto).order_by(Produto.codigo, Contagem.lote).all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Estoque"
        
        # Estilos
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        subtotal_font = Font(bold=True)
        subtotal_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalho
        ws['A1'] = "Relatório de Estoque"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = center_alignment
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = center_alignment
        
        ws['A4'] = "Relatório de Estoque - Detalhado por Lote"
        ws['A4'].font = Font(bold=True, size=14)
        ws.merge_cells('A4:F4')
        
        # Cabeçalhos da tabela
        headers = ['Código', 'Nome do Produto', 'Lote', 'Validade', 'Qtd', 'Cadastro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Organizar dados por produto
        produtos_dict = {}
        for produto in produtos_data:
            produtos_dict[produto.codigo] = {
                'nome': produto.nome,
                'contagens': [],
                'subtotal': 0
            }
        
        # Adicionar contagens aos produtos
        for contagem in contagens_data:
            if contagem.codigo in produtos_dict:
                validade = f"{contagem.validade_mes:02d}/{contagem.validade_ano}"
                cadastro = '-'  # Campo não disponível no modelo atual
                
                produtos_dict[contagem.codigo]['contagens'].append({
                    'lote': contagem.lote,
                    'validade': validade,
                    'quantidade': contagem.quantidade,
                    'cadastro': cadastro
                })
                produtos_dict[contagem.codigo]['subtotal'] += contagem.quantidade
        
        total_geral = 0
        row = 7
        
        # Adicionar dados à planilha
        for codigo in sorted(produtos_dict.keys()):
            produto = produtos_dict[codigo]
            
            if produto['contagens']:
                # Produto com contagens
                for contagem in produto['contagens']:
                    ws.cell(row=row, column=1, value=codigo).border = border
                    ws.cell(row=row, column=2, value=produto['nome']).border = border
                    ws.cell(row=row, column=3, value=contagem['lote']).border = border
                    ws.cell(row=row, column=4, value=contagem['validade']).border = border
                    ws.cell(row=row, column=4).alignment = center_alignment
                    ws.cell(row=row, column=5, value=contagem['quantidade']).border = border
                    ws.cell(row=row, column=5).alignment = center_alignment
                    ws.cell(row=row, column=6, value=contagem['cadastro']).border = border
                    ws.cell(row=row, column=6).alignment = center_alignment
                    row += 1
                
                # Linha de subtotal
                ws.cell(row=row, column=1, value=f'Subtotal {codigo}:').border = border
                ws.cell(row=row, column=1).font = subtotal_font
                ws.cell(row=row, column=1).fill = subtotal_fill
                for col in range(2, 5):
                    ws.cell(row=row, column=col, value='').border = border
                    ws.cell(row=row, column=col).fill = subtotal_fill
                ws.cell(row=row, column=5, value=produto['subtotal']).border = border
                ws.cell(row=row, column=5).font = subtotal_font
                ws.cell(row=row, column=5).fill = subtotal_fill
                ws.cell(row=row, column=5).alignment = center_alignment
                ws.cell(row=row, column=6, value='').border = border
                ws.cell(row=row, column=6).fill = subtotal_fill
                total_geral += produto['subtotal']
                row += 1
            else:
                # Produto sem contagens (estoque zerado)
                ws.cell(row=row, column=1, value=codigo).border = border
                ws.cell(row=row, column=2, value=produto['nome']).border = border
                ws.cell(row=row, column=3, value='-').border = border
                ws.cell(row=row, column=4, value='-').border = border
                ws.cell(row=row, column=4).alignment = center_alignment
                ws.cell(row=row, column=5, value=0).border = border
                ws.cell(row=row, column=5).alignment = center_alignment
                ws.cell(row=row, column=6, value='-').border = border
                ws.cell(row=row, column=6).alignment = center_alignment
                row += 1
                
                # Linha de subtotal para produto zerado
                ws.cell(row=row, column=1, value=f'Subtotal {codigo}:').border = border
                ws.cell(row=row, column=1).font = subtotal_font
                ws.cell(row=row, column=1).fill = subtotal_fill
                for col in range(2, 5):
                    ws.cell(row=row, column=col, value='').border = border
                    ws.cell(row=row, column=col).fill = subtotal_fill
                ws.cell(row=row, column=5, value=0).border = border
                ws.cell(row=row, column=5).font = subtotal_font
                ws.cell(row=row, column=5).fill = subtotal_fill
                ws.cell(row=row, column=5).alignment = center_alignment
                ws.cell(row=row, column=6, value='').border = border
                ws.cell(row=row, column=6).fill = subtotal_fill
                row += 1
        
        # Linha de total geral
        ws.cell(row=row, column=1, value='TOTAL GERAL:').border = border
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        for col in range(2, 5):
            ws.cell(row=row, column=col, value='').border = border
            ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=5, value=total_geral).border = border
        ws.cell(row=row, column=5).font = total_font
        ws.cell(row=row, column=5).fill = total_fill
        ws.cell(row=row, column=5).alignment = center_alignment
        ws.cell(row=row, column=6, value='').border = border
        ws.cell(row=row, column=6).fill = total_fill
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y-%m-%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar Excel: {str(e)}'}), 500

